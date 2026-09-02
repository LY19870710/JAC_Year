#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
server.py — JAC_Year Web 界面 v0.4.2
Python 标准库实现，无需 npm/Express
访问: http://localhost:3000

v0.4.2 改动：导出按钮(XLSX/CSV/JSON/EndNote)、单篇RIS📥
  - 通讯作者-邮箱对应关系（JSON格式，每人独立显示）
  - 机构名称 tooltip（hover 显示完整原文）
  - 机构提取优化（University/Institute/Academy/Laboratory 关键词定位）
  - funding 字段展示（可折叠）
"""
import sqlite3, json, sys, os, re, logging
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
sys.stderr.reconfigure(encoding='utf-8', errors='replace')
from pathlib import Path
from http.server import HTTPServer, BaseHTTPRequestHandler
from urllib.parse import urlparse, parse_qs, quote
from datetime import datetime

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler('server.log', encoding='utf-8')
    ]
)
logger = logging.getLogger(__name__)

DB_PATH = Path(__file__).parent.parent / "jac_articles.db"
PORT    = int(os.environ.get("PORT", 3000))

# 用户认证配置
AUTH_ENABLED = os.environ.get("AUTH_ENABLED", "false").lower() == "true"
AUTH_USER = os.environ.get("AUTH_USER", "admin")
AUTH_PASS = os.environ.get("AUTH_PASS", "password")

ARTICLE_TYPES = [
    "Research Article", "Review", "Editorial",
    "Erratum", "Perspective", "Rapid Communication"
]

AREA_COLORS = {
    1:  "#dc2626", 2:  "#ea580c", 3:  "#d97706",
    4:  "#65a30d", 5:  "#0891b2", 6:  "#7c3aed",
    7:  "#db2777", 8:  "#0284c7", 9:  "#16a34a",
    10: "#9333ea", 11: "#c2410c", 12: "#0369a1", 0: "#6b7280",
}

# ─── DB helpers ──────────────────────────────────────────────

def get_conn():
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    return conn

def get_years(conn):
    return [r[0] for r in conn.execute("SELECT DISTINCT year FROM articles ORDER BY year DESC")]

def get_areas(conn):
    return conn.execute(
        "SELECT DISTINCT research_area_id, research_area, research_area_zh, COUNT(*) as cnt "
        "FROM articles GROUP BY research_area_id ORDER BY cnt DESC"
    ).fetchall()

def extract_institution(aff_item: str) -> tuple[str, str]:
    """
    从单条机构原文中提取机构名。
    返回 (short_name, original) 元组。
    
    规则：提取城市/邮编前面的最后一个机构名称
    例如: "School of Materials Science and Engineering, Gyeongsang National University, Gyeongsangnam-do 52828, Republic of Korea"
    返回: "Gyeongsang National University"
    """
    original = aff_item.strip()
    
    # 去掉开头的数字编号
    cleaned = re.sub(r'^[\d†*]+', '', original).strip()
    if not cleaned:
        return ('', original)
    
    # 跳过噪音行
    if 'contributed equally' in cleaned.lower():
        # 移除"contributed equally"部分，继续处理
        cleaned = re.sub(r'contributed equally,?\s*', '', cleaned, flags=re.IGNORECASE).strip()
        if not cleaned:
            return ('', original)
    if cleaned.startswith('†') or cleaned.startswith('*'):
        return ('', original)
    
    # 按逗号分隔
    parts = [p.strip() for p in cleaned.split(',')]
    
    # 国家名列表
    COUNTRIES = {
        'china', 'usa', 'u.s.a.', 'uk', 'japan', 'korea', 'germany', 'france',
        'republic of korea', 'south korea', 'australia', 'canada', 'india',
        'italy', 'spain', 'russia', 'netherlands', 'brazil', 'mexico',
        'thailand', 'singapore', 'malaysia', 'vietnam', 'indonesia',
        'philippines', 'pakistan', 'bangladesh', 'sri lanka', 'iran',
        'turkey', 'egypt', 'south africa', 'nigeria', 'kenya',
        'chile', 'argentina', 'colombia', 'peru', 'venezuela',
        'new zealand', 'austria', 'belgium', 'denmark', 'finland',
        'greece', 'ireland', 'norway', 'poland', 'portugal',
        'sweden', 'switzerland', 'czech republic', 'hungary', 'romania',
        'ukraine', 'belarus', 'serbia', 'croatia', 'slovenia',
        'estonia', 'latvia', 'lithuania', 'iceland', 'luxembourg',
    }
    
    # 城市名列表（常见的）
    CITIES = {
        'beijing', 'shanghai', 'guangzhou', 'shenzhen', 'chengdu', 'wuhan',
        'xi\'an', 'xian', 'hangzhou', 'nanjing', 'harbin', 'jinan', 'changsha',
        'daegu', 'seoul', 'tokyo', 'osaka', 'bangkok', 'singapore', 'new york',
        'los angeles', 'london', 'paris', 'berlin', 'moscow', 'toronto',
        'sydney', 'melbourne', 'delhi', 'mumbai', 'bangkok', 'zhengzhou',
        'kaifeng', 'xiamen', 'dalian', 'qingdao', 'jinan', 'hefei',
    }
    
    def is_country(p: str) -> bool:
        p_lower = p.lower().strip()
        # 精确匹配
        if p_lower in COUNTRIES:
            return True
        # 检查是否包含国家名（如 "China Academy of Sciences"）
        for country in COUNTRIES:
            if country in p_lower:
                return True
        return False
    
    def is_city(p: str) -> bool:
        """检查是否是城市名"""
        return p.lower().strip() in CITIES
    
    def is_city_zip(p: str) -> bool:
        """检查是否是城市+邮编格式（如 "Kaifeng 475004"）"""
        return bool(re.search(r'\d{4,}', p))
    
    def is_address(p: str) -> bool:
        """检查是否是地址部分（城市、邮编、国家）"""
        p_lower = p.lower().strip()
        if is_country(p):
            return True
        if is_city(p):
            return True
        if is_city_zip(p):
            return True
        # 检查是否包含邮编格式
        if re.search(r'\b\d{4,6}\b', p):
            return True
        return False
    
    def is_noise(p: str) -> bool:
        """检查是否是噪音（如 "Ltd.", "Inc.", 单个字母等）"""
        p_lower = p.lower().strip()
        # 太短的通常是噪音
        if len(p_lower) < 3:
            return True
        # 单独的公司后缀
        if p_lower in ('ltd.', 'ltd', 'inc.', 'inc', 'co.', 'co', 'corp.', 'corp', 'gmbh', 'ag', 'sa', 'bv'):
            return True
        return False
    
    # 从后往前找，找到城市/地址前的最后一个机构名称
    for i in range(len(parts) - 1, -1, -1):
        p = parts[i].strip()
        if not p:
            continue
        if is_address(p):
            continue
        if is_noise(p):
            continue
        # 找到了城市前面的最后一个机构名称
        # 清理末尾的噪音
        p_clean = re.sub(r'\s*,?\s*(Ltd\.|Inc\.|Co\.|Corp\.)\s*$', '', p, flags=re.IGNORECASE)
        return (p_clean, original)
    
    # 如果没找到，返回第一个非空、非噪音部分
    for p in parts:
        p_clean = p.strip()
        if p_clean and not is_noise(p_clean):
            return (p_clean, original)
    
    return ('', original)


def query_articles(conn, filters: dict, limit=500, offset=0):
    sql = "SELECT * FROM articles WHERE 1=1 AND type != 'Technical Paper'"
    params = []
    if filters.get("year"):
        sql += " AND year=?"; params.append(int(filters["year"]))
    if filters.get("type"):
        sql += " AND type=?"; params.append(filters["type"])
    if filters.get("area"):
        sql += " AND research_area_id=?"; params.append(int(filters["area"]))
    if filters.get("author"):
        sql += " AND authors LIKE ?"; params.append(f"%{filters['author']}%")
    if filters.get("keyword"):
        sql += " AND (title LIKE ? OR authors LIKE ?)"; params += [f"%{filters['keyword']}%"]*2
    sql += " ORDER BY year DESC, volume DESC, issue DESC, doi LIMIT ? OFFSET ?"
    params.extend([limit, offset])
    return conn.execute(sql, params).fetchall()


def count_articles(conn, filters: dict):
    sql = "SELECT COUNT(*) FROM articles WHERE 1=1 AND type != 'Technical Paper'"
    params = []
    if filters.get("year"):
        sql += " AND year=?"; params.append(int(filters["year"]))
    if filters.get("type"):
        sql += " AND type=?"; params.append(filters["type"])
    if filters.get("area"):
        sql += " AND research_area_id=?"; params.append(int(filters["area"]))
    if filters.get("author"):
        sql += " AND authors LIKE ?"; params.append(f"%{filters['author']}%")
    if filters.get("keyword"):
        sql += " AND (title LIKE ? OR authors LIKE ?)"; params += [f"%{filters['keyword']}%"]*2
    return conn.execute(sql, params).fetchone()[0]

def get_stats(conn):
    total = conn.execute("SELECT COUNT(*) FROM articles").fetchone()[0]
    by_year  = conn.execute("SELECT year, COUNT(*) FROM articles GROUP BY year ORDER BY year DESC").fetchall()
    by_type  = conn.execute("SELECT type, COUNT(*) FROM articles GROUP BY type ORDER BY COUNT(*) DESC").fetchall()
    by_area  = conn.execute(
        "SELECT research_area_id, research_area, research_area_zh, COUNT(*) as cnt "
        "FROM articles GROUP BY research_area_id ORDER BY cnt DESC"
    ).fetchall()
    return {"total": total, "by_year": by_year, "by_type": by_type, "by_area": by_area}


# ─── HTML templates ──────────────────────────────────────────

CSS = """
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:-apple-system,BlinkMacSystemFont,'PingFang SC','Segoe UI',sans-serif;background:#f5f5f5;color:#1c1917}
a{color:#ea580c;text-decoration:none}a:hover{text-decoration:underline}
.wrap{max-width:1300px;margin:0 auto;padding:16px}
header{background:linear-gradient(135deg,#ea580c,#9a3412);color:#fff;padding:18px 24px;margin-bottom:16px}
header h1{font-size:1.3rem;font-weight:700}
header p{font-size:.85rem;opacity:.85;margin-top:4px}
nav{background:#fff;padding:10px 20px;border-radius:8px;margin-bottom:16px;display:flex;gap:20px;box-shadow:0 1px 3px rgba(0,0,0,.08)}
nav a{color:#57534e;font-size:.9rem;padding:4px 0}nav a.active,nav a:hover{color:#ea580c;border-bottom:2px solid #ea580c}
.card{background:#fff;border-radius:10px;padding:20px;margin-bottom:16px;box-shadow:0 1px 3px rgba(0,0,0,.08)}
.card h2{font-size:1rem;font-weight:700;margin-bottom:14px;color:#1c1917}
.form-row{display:flex;flex-wrap:wrap;gap:12px;align-items:flex-end}
.fg{display:flex;flex-direction:column;gap:4px;min-width:140px}
.fg label{font-size:.8rem;color:#78716c;font-weight:600}
.fg input,.fg select{padding:7px 10px;border:1px solid #e7e5e4;border-radius:6px;font-size:.875rem;background:#fff}
.fg input:focus,.fg select:focus{outline:none;border-color:#ea580c}
.btn{background:#ea580c;color:#fff;border:none;padding:8px 20px;border-radius:6px;cursor:pointer;font-size:.875rem;font-weight:600;white-space:nowrap}
.btn:hover{background:#c2410c}
.btn-sm{padding:4px 10px;font-size:.75rem;border-radius:4px}
.count{font-size:.875rem;color:#78716c;margin-bottom:12px}
table{width:100%;border-collapse:collapse;font-size:.85rem}
th{background:#fff7ed;padding:10px 12px;text-align:left;font-weight:700;color:#9a3412;border-bottom:2px solid #fed7aa;white-space:nowrap}
td{padding:9px 12px;border-bottom:1px solid #f5f5f4;vertical-align:top}
tr:hover td{background:#fff7ed}
.tag{display:inline-block;padding:2px 8px;border-radius:12px;font-size:.72rem;font-weight:700;white-space:nowrap}
.tag-type{background:#e0e7ff;color:#3730a3}
.area-badge{display:inline-block;padding:2px 8px;border-radius:12px;font-size:.72rem;font-weight:600;color:#fff;white-space:nowrap}
.title-cell{max-width:380px}
.authors-cell{max-width:200px;color:#57534e;font-size:.8rem}
.doi-link{font-family:monospace;font-size:.78rem;color:#ea580c}
.stats-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(220px,1fr));gap:16px;margin-bottom:16px}
.stat-box{background:#fff7ed;border-radius:10px;padding:20px;text-align:center;border:1px solid #fed7aa}
.stat-num{font-size:2.2rem;font-weight:800;color:#ea580c}
.stat-lbl{font-size:.85rem;color:#78716c;margin-top:4px}
.bar-row{display:flex;align-items:center;gap:10px;margin-bottom:8px}
.bar-label{width:260px;font-size:.82rem;color:#1c1917;flex-shrink:0;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.bar-track{flex:1;background:#f5f5f4;border-radius:4px;height:18px;overflow:hidden}
.bar-fill{height:100%;border-radius:4px;transition:width .3s}
.bar-count{width:36px;text-align:right;font-size:.82rem;font-weight:700;color:#78716c;flex-shrink:0}
.empty{text-align:center;padding:40px;color:#a8a29e}
/* tooltip */
.tip{position:relative;cursor:help;border-bottom:1px dashed #d1c4b0}
.tip:hover::after{content:attr(data-tip);position:absolute;left:0;top:100%;z-index:99;background:#1c1917;color:#fff;padding:6px 10px;border-radius:6px;font-size:.75rem;white-space:pre-wrap;max-width:360px;line-height:1.5;box-shadow:0 4px 12px rgba(0,0,0,.25);pointer-events:none;margin-top:4px}
/* funding */
.funding-toggle{font-size:.72rem;color:#ea580c;cursor:pointer;user-select:none;margin-top:4px;display:inline-block}
.funding-toggle:hover{text-decoration:underline}
.funding-body{display:none;font-size:.75rem;color:#57534e;margin-top:4px;padding:6px 8px;background:#fff7ed;border-radius:4px;border-left:3px solid #fed7aa;line-height:1.5}
/* corr authors */
.corr-list{margin-top:4px;font-size:.78rem}
.corr-item{display:flex;align-items:baseline;gap:4px;line-height:1.6}
.corr-name{color:#1c1917;font-weight:600;white-space:nowrap}
.corr-email{color:#ea580c;font-size:.72rem}
"""

def layout(title, body, active="search"):
    nav_links = [
        ("search", "/", "Search"),
        ("stats",  "/stats", "Stats"),
    ]
    nav_html = "".join(
        f'<a href="{href}" class="{("active" if k==active else "")}">{label}</a>'
        for k, href, label in nav_links
    )
    return f"""<!DOCTYPE html>
<html lang="zh-CN">
<head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{title} — JAC Year</title>
<style>{CSS}</style>
</head>
<body>
<header>
  <div class="wrap">
    <h1>Journal of Advanced Ceramics — Year Manager</h1>
    <p>JAC_Year v0.4.2</p>
  </div>
</header>
<div class="wrap">
  <nav>{nav_html}</nav>
  {body}
</div>
</body></html>"""


def esc(s):
    if s is None: return ""
    return str(s).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;").replace('"',"&quot;")


def render_search(conn, filters=None, articles=None, total=0, page=1, per_page=50):
    filters = filters or {}
    years = get_years(conn)
    areas = get_areas(conn)

    year_opts = '<option value="">All Years</option>' + "".join(
        f'<option value="{y}" {"selected" if str(filters.get("year",""))==str(y) else ""}>{y}</option>'
        for y in years
    )
    type_opts = '<option value="">All Types</option>' + "".join(
        f'<option value="{t}" {"selected" if filters.get("type","")==t else ""}>{t}</option>'
        for t in ARTICLE_TYPES
    )
    area_opts = '<option value="">All Areas</option>' + "".join(
        f'<option value="{r["research_area_id"]}" {"selected" if str(filters.get("area",""))==str(r["research_area_id"]) else ""}>{r["research_area_zh"]} ({r["cnt"]})</option>'
        for r in areas
    )

    form = f"""<div class="card">
  <h2>Search文章</h2>
  <form method="GET" action="/search">
    <div class="form-row">
      <div class="fg"><label>Year</label><select name="year">{year_opts}</select></div>
      <div class="fg"><label>文章Type</label><select name="type">{type_opts}</select></div>
      <div class="fg" style="min-width:200px"><label>Research Area</label><select name="area">{area_opts}</select></div>
      <div class="fg"><label>Author</label><input name="author" value="{esc(filters.get('author',''))}" placeholder="Author姓名"></div>
      <div class="fg"><label>Keyword</label><input name="keyword" value="{esc(filters.get('keyword',''))}" placeholder="Title/AuthorKeyword"></div>
      <div class="fg"><label>&nbsp;</label><button class="btn" type="submit">Search</button></div>
    </div>
  </form>
</div>"""

    if articles is None:
        body = form + '<div class="card"><p class="empty">输入条件后点击Search</p></div>'
        return layout("Search", body, "search")

    rows = ""
    for a in articles:
        color = AREA_COLORS.get(a["research_area_id"], "#6b7280")

        # ── 通讯作者（JSON格式，每人独立一行）──
        corr_json_str = a["corresponding_json"] if "corresponding_json" in a.keys() else "[]"
        try:
            corr_list = json.loads(corr_json_str or "[]")
        except Exception:
            corr_list = []

        if corr_list:
            items_html = ""
            for c in corr_list:
                raw_name = c.get("name", "")
                # 剥离HTML标签（sciopen有时返回<a>标签）
                name = esc(re.sub(r'<[^>]+>', '', raw_name).strip()) if raw_name else ""
                email = esc(c.get("email", ""))
                email_html = f'<a class="corr-email" href="mailto:{email}">{email}</a>' if email else ""
                items_html += f'<div class="corr-item"><span class="corr-name">✉ {name}</span>{email_html}</div>'
            corr_html = f'<div class="corr-list">{items_html}</div>'
        else:
            corr_html = ""

        # ── 机构提取（优化版 + tooltip）──
        aff_raw = a["affiliations"] or ""
        if aff_raw:
            aff_parts = []
            seen_short = set()
            for aff_item in aff_raw.split(";"):
                short, original = extract_institution(aff_item)
                if not short:
                    continue
                if short in seen_short:
                    continue
                seen_short.add(short)
                # tooltip 显示原文（换行转义）
                import html as _html_mod
                tip_text = esc(_html_mod.unescape(original)).replace('\n', ' ')
                aff_parts.append(f'<span class="tip" data-tip="{tip_text}">{esc(short)}</span>')
            aff_html = "; ".join(aff_parts) if aff_parts else '<span style="color:#d1c4b0">—</span>'
        else:
            aff_html = '<span style="color:#d1c4b0">—</span>'

        # ── Funding（可折叠）──
        funding = (a["funding"] if "funding" in a.keys() else "") or ""
        if funding.strip():
            uid = f"f{a['id']}"
            funding_html = (
                f'<div class="funding-toggle" onclick="'
                f'var b=document.getElementById(\'{uid}\');'
                f'b.style.display=b.style.display===\'block\'?\'none\':\'block\'">'
                f'▶ Funding</div>'
                f'<div class="funding-body" id="{uid}">{esc(funding)}</div>'
            )
        else:
            funding_html = ""

        # ── Volume/Issue ──
        vol_issue = f"Vol.{a['volume'] or '?'} Issue {a['issue'] or '?'}"

        rows += f"""<tr>
  <td style="white-space:nowrap;font-size:.82rem">{vol_issue}</td>
  <td><span class="tag tag-type">{esc(a['type'])}</span></td>
  <td class="title-cell">
    <a href="{esc(a['url'])}" target="_blank">{esc(a['title'])}</a>
    {funding_html}
  </td>
  <td class="authors-cell">{esc(a['authors'])}{corr_html}</td>
  <td style="font-size:.78rem;color:#57534e;max-width:240px">{aff_html}</td>
  <td><span class="area-badge" style="background:{color}">{esc(a['research_area_zh'])}</span></td>
  <td style="white-space:nowrap"><a class="doi-link" href="{esc(a['url'])}" target="_blank">{esc(a['doi'])}</a> <a href="/api/export-ris?ids={a['id']}" style="color:#ea580c;text-decoration:none;font-size:.85rem" target="_blank" title="导出RIS">📥</a></td>
</tr>"""

    # ── 导出按钮 ──
    export_qs = ""
    for k in ("year", "type", "area", "author", "keyword"):
        if filters.get(k):
            v = filters[k]
            if k == "area":
                k = "area"
            export_qs += f"&{k}={quote(str(v))}"

    export_btns = f"""<div style="display:flex;gap:8px;align-items:center;margin:8px 0">
  <span style="color:#666;font-size:.85rem">导出当前结果：</span>
  <a href="/api/export?format=xlsx{export_qs}" class="btn" style="background:#16a34a;color:#fff;text-decoration:none;font-size:.82rem">XLSX</a>
  <a href="/api/export?format=csv{export_qs}" class="btn" style="background:#2563eb;color:#fff;text-decoration:none;font-size:.82rem">CSV</a>
  <a href="/api/export?format=json{export_qs}" class="btn" style="background:#7c3aed;color:#fff;text-decoration:none;font-size:.82rem">JSON</a>
  <a href="/api/export-ris?{export_qs[1:]}" class="btn" style="background:#ea580c;color:#fff;text-decoration:none;font-size:.82rem" title="导出RIS格式，可导入EndNote">EndNote</a>
</div>"""

    table = f"""<div class="card">
  <p class="count">Found <strong>{total}</strong> articles (showing {len(articles)} of {total})</p>
  {export_btns}
  <table>
    <thead><tr><th>Issue</th><th>Type</th><th>Title</th><th>Authors / Corr.</th><th>Affiliations</th><th>Research Area</th><th>DOI</th></tr></thead>
    <tbody>{rows if rows else '<tr><td colspan="7" class="empty">No results</td></tr>'}</tbody>
  </table>
</div>"""

    # 分页控件
    total_pages = (total + per_page - 1) // per_page
    if total_pages > 1:
        pagination_html = '<div style="display:flex;justify-content:center;gap:8px;margin:16px 0">'
        if page > 1:
            pagination_html += f'<a href="?page={page-1}{export_qs}" class="btn" style="background:#6b7280;color:#fff;text-decoration:none">上一页</a>'
        
        # 显示页码
        for p in range(1, min(total_pages + 1, 11)):
            if p == page:
                pagination_html += f'<span class="btn" style="background:#ea580c;color:#fff">{p}</span>'
            else:
                pagination_html += f'<a href="?page={p}{export_qs}" class="btn" style="background:#f5f5f4;color:#1c1917;text-decoration:none">{p}</a>'
        
        if page < total_pages:
            pagination_html += f'<a href="?page={page+1}{export_qs}" class="btn" style="background:#6b7280;color:#fff;text-decoration:none">下一页</a>'
        
        pagination_html += '</div>'
        table += pagination_html

    body = form + table
    return layout(f"Search Results ({total})", body, "search")


def render_stats(conn):
    stats = get_stats(conn)
    total = stats["total"]
    years = stats["by_year"]
    types = stats["by_type"]
    areas = stats["by_area"]

    # 顶部数字卡
    year_list = ", ".join(str(r[0]) for r in years)
    top_cards = f"""<div class="stats-grid">
  <div class="stat-box"><div class="stat-num">{total}</div><div class="stat-lbl">总文章数</div></div>
  <div class="stat-box"><div class="stat-num">{len(years)}</div><div class="stat-lbl">覆盖Year<br><small style="font-size:.75rem;color:#a16207">{year_list}</small></div></div>
  <div class="stat-box"><div class="stat-num">{len(areas)}</div><div class="stat-lbl">Research Area</div></div>
</div>"""

    # Year分布
    max_yr = max((r[1] for r in years), default=1)
    yr_bars = "".join(
        f'<div class="bar-row"><div class="bar-label">{r[0]}</div>'
        f'<div class="bar-track"><div class="bar-fill" style="width:{r[1]/max_yr*100:.0f}%;background:#ea580c"></div></div>'
        f'<div class="bar-count">{r[1]}</div></div>'
        for r in years
    )

    # Research Area分布
    max_area = max((r["cnt"] for r in areas), default=1)
    area_bars = "".join(
        f'<div class="bar-row">'
        f'<div class="bar-label" title="{esc(r["research_area"])}">{esc(r["research_area_zh"])}</div>'
        f'<div class="bar-track"><div class="bar-fill" style="width:{r["cnt"]/max_area*100:.0f}%;background:{AREA_COLORS.get(r["research_area_id"],"#6b7280")}"></div></div>'
        f'<div class="bar-count">{r["cnt"]}</div></div>'
        for r in areas
    )

    # 文章Type
    max_type = max((r[1] for r in types), default=1)
    type_bars = "".join(
        f'<div class="bar-row"><div class="bar-label">{esc(r[0])}</div>'
        f'<div class="bar-track"><div class="bar-fill" style="width:{r[1]/max_type*100:.0f}%;background:#7c3aed"></div></div>'
        f'<div class="bar-count">{r[1]}</div></div>'
        for r in types
    )

    body = top_cards + f"""
<div style="display:grid;grid-template-columns:1fr 1fr;gap:16px">
  <div class="card"><h2>Year分布</h2><div style="margin-top:12px">{yr_bars}</div></div>
  <div class="card"><h2>文章Type</h2><div style="margin-top:12px">{type_bars}</div></div>
</div>
<div class="card"><h2>Research Area分布</h2><div style="margin-top:12px">{area_bars}</div></div>
"""
    return layout("Stats", body, "stats")


# ─── HTTP Handler ─────────────────────────────────────────────

class Handler(BaseHTTPRequestHandler):
    def log_message(self, fmt, *args):
        # 使用logging模块记录请求日志
        logger.info(f"{self.client_address[0]} - {fmt % args}")

    def check_auth(self):
        """检查HTTP基本认证"""
        if not AUTH_ENABLED:
            return True
        
        auth_header = self.headers.get('Authorization')
        if not auth_header or not auth_header.startswith('Basic '):
            return False
        
        import base64
        try:
            auth_decoded = base64.b64decode(auth_header[6:]).decode('utf-8')
            username, password = auth_decoded.split(':', 1)
            return username == AUTH_USER and password == AUTH_PASS
        except:
            return False

    def require_auth(self):
        """要求认证，失败返回401"""
        if not self.check_auth():
            self.send_response(401)
            self.send_header('WWW-Authenticate', 'Basic realm="JAC_Year"')
            self.send_header('Content-type', 'text/html; charset=utf-8')
            self.end_headers()
            self.wfile.write(b'401 Unauthorized')
            return False
        return True

    def send_html(self, html, code=200):
        b = html.encode("utf-8")
        self.send_response(code)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", len(b))
        self.end_headers()
        self.wfile.write(b)

    def send_json(self, data):
        b = json.dumps(data, ensure_ascii=False, default=str).encode("utf-8")
        self.send_response(200)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", len(b))
        self.end_headers()
        self.wfile.write(b)

    def send_file(self, data: bytes, filename: str, content_type: str):
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
        self.send_header("Content-Length", len(data))
        self.end_headers()
        self.wfile.write(data)

    def _handle_export_ris(self, conn, params):
        """导出 RIS 格式（EndNote 导入用）"""
        ids = params.get("ids", "")
        if ids:
            # 单篇/多篇按 id
            id_list = [int(x) for x in ids.split(",") if x.strip().isdigit()]
            placeholders = ",".join("?" * len(id_list))
            query = f"SELECT doi, title, authors, volume, issue, year, abstract, url FROM articles WHERE id IN ({placeholders})"
            rows = conn.execute(query, id_list).fetchall()
        else:
            # 按筛选条件
            rows = query_articles(conn, params, limit=2000)
            rows = [dict(r) for r in rows]

        lines = []
        for row in rows:
            doi    = row["doi"]    if isinstance(row, dict) else row[0]
            title  = row["title"]  if isinstance(row, dict) else row[1]
            authors= row["authors"]if isinstance(row, dict) else row[2]
            volume = row["volume"] if isinstance(row, dict) else row[3]
            issue  = row["issue"]  if isinstance(row, dict) else row[4]
            year   = row["year"]   if isinstance(row, dict) else row[5]
            abstract=row["abstract"]if isinstance(row, dict)else row[6]
            url    = row["url"]    if isinstance(row, dict) else row[7]

            # SP 从 DOI 提取文章编号
            sp = ""
            m = re.search(r'JAC\.\d{4}\.(\d+)', doi or "")
            if m:
                sp = m.group(1)
            # PY 从 DOI 提取年份
            py = ""
            m2 = re.search(r'JAC\.(\d{4})', doi or "")
            if m2:
                py = m2.group(1)

            lines.append("TY  - JOUR")
            # AU - 每人一行 Last, First
            if authors:
                for au in authors.split(", "):
                    au = au.strip()
                    if not au or au == "[...]":
                        continue
                    parts = au.rsplit(" ", 1)
                    if len(parts) == 2:
                        lines.append(f"AU  - {parts[1]}, {parts[0]}")
                    else:
                        lines.append(f"AU  - {au}")
            if py:
                lines.append(f"PY  - {py}")
            # 清理 HTML
            clean_title = re.sub(r'<[^>]+>', '', title or "").strip()
            lines.append(f"TI  - {clean_title}")
            lines.append("JO  - Journal of Advanced Ceramics")
            lines.append("SN  - 2226-4108")
            if sp:
                lines.append(f"SP  - {sp}")
            if volume:
                lines.append(f"VL  - {volume}")
            if issue:
                lines.append(f"IS  - {issue}")
            if abstract:
                import html as _html
                ab = re.sub(r'<[^>]+>', '', abstract)
                ab = _html.unescape(ab)
                ab = re.sub(r'\\u([0-9a-fA-F]{4})', lambda m: chr(int(m.group(1), 16)), ab)
                ab = re.sub(r'\s+', ' ', ab).strip()
                lines.append(f"AB  - {ab}")
            if url:
                lines.append(f"UR  - {url}")
            if doi:
                lines.append(f"DO  - {doi}")
            lines.append("")  # 空行分隔

        data = "\n".join(lines).encode("utf-8")
        self.send_file(data, "JAC_export.ris", "application/x-research-info-systems; charset=utf-8")

    def _handle_export(self, conn, params):
        """导出 XLSX/CSV/JSON"""
        fmt = params.get("format", "json")
        articles = query_articles(conn, params, limit=2000)
        articles = [dict(a) for a in articles]

        # ── 清理 & 精选导出列 ──
        import html as _html
        def clean(s):
            """清理 HTML 标签和实体"""
            if not s: return ""
            s = re.sub(r'<[^>]+>', '', str(s))
            s = _html.unescape(s)
            s = re.sub(r'\\u([0-9a-fA-F]{4})', lambda m: chr(int(m.group(1), 16)), s)
            s = re.sub(r'\s+', ' ', s).strip()
            return s

        def parse_corr_json(raw):
            """从 corresponding_json 提取通讯作者信息"""
            try:
                lst = json.loads(raw or "[]")
                if not lst: return "", "", ""
                names = "; ".join(c.get("name", "") for c in lst if c.get("name"))
                emails = "; ".join(c.get("email", "") for c in lst if c.get("email"))
                insts = "; ".join(c.get("institution", "") for c in lst if c.get("institution"))
                return names, emails, insts
            except Exception:
                return "", "", ""

        EXPORT_COLS = [
            ("DOI", lambda a: a.get("doi", "")),
            ("Title", lambda a: clean(a.get("title", ""))),
            ("Authors", lambda a: clean(a.get("authors", ""))),
            ("Corresponding Author", lambda a: parse_corr_json(a.get("corresponding_json", ""))[0]),
            ("Corresponding Email", lambda a: parse_corr_json(a.get("corresponding_json", ""))[1]),
            ("Corresponding Institution", lambda a: parse_corr_json(a.get("corresponding_json", ""))[2]),
            ("Affiliations", lambda a: clean(a.get("affiliations", ""))),
            ("Year", lambda a: a.get("year", "")),
            ("Volume", lambda a: a.get("volume", "")),
            ("Issue", lambda a: a.get("issue", "")),
            ("Type", lambda a: a.get("type", "")),
            ("Research Area", lambda a: a.get("research_area_zh", "")),
            ("Keywords", lambda a: clean(a.get("keywords", ""))),
            ("Abstract", lambda a: clean(a.get("abstract", ""))),
            ("Funding", lambda a: clean(a.get("funding", ""))),
            ("Citation", lambda a: clean(a.get("citation", ""))),
            ("Received", lambda a: a.get("received_date", "")),
            ("Accepted", lambda a: a.get("accepted_date", "")),
            ("Published", lambda a: a.get("published_date", "")),
            ("URL", lambda a: a.get("url", "")),
        ]
        col_names = [c[0] for c in EXPORT_COLS]
        clean_articles = []
        for a in articles:
            row = {c[0]: c[1](a) for c in EXPORT_COLS}
            clean_articles.append(row)

        if fmt == "json":
            data = json.dumps(clean_articles, ensure_ascii=False, default=str).encode("utf-8")
            self.send_file(data, "JAC_export.json", "application/json; charset=utf-8")
        elif fmt == "csv":
            import csv, io as _io
            buf = _io.StringIO()
            if clean_articles:
                w = csv.DictWriter(buf, fieldnames=col_names)
                w.writeheader()
                w.writerows(clean_articles)
            data = buf.getvalue().encode("utf-8-sig")
            self.send_file(data, "JAC_export.csv", "text/csv; charset=utf-8")
        elif fmt == "xlsx":
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
            wb = Workbook()
            ws = wb.active
            ws.title = "JAC Articles"
            # 表头加粗
            header_font = Font(bold=True)
            wrap_align = Alignment(wrap_text=True, vertical="top")
            for ci, name in enumerate(col_names, 1):
                cell = ws.cell(1, ci, name)
                cell.font = header_font
            for ri, row in enumerate(clean_articles, 2):
                for ci, name in enumerate(col_names, 1):
                    cell = ws.cell(ri, ci, row.get(name, ""))
                    cell.alignment = wrap_align
            # 自动列宽
            for ci, name in enumerate(col_names, 1):
                cell_lens = [len(str(ws.cell(r, ci).value or "")) for r in range(2, min(20, ws.max_row + 1))]
                max_len = max(len(str(name)), max(cell_lens, default=0))
                ws.column_dimensions[ws.cell(1, ci).column_letter].width = min(max_len + 2, 60)
            # DOI列超链接
            doi_col = col_names.index("DOI") + 1
            url_col = col_names.index("URL") + 1
            for r in range(2, ws.max_row + 1):
                url_val = ws.cell(r, url_col).value
                if url_val:
                    ws.cell(r, doi_col).hyperlink = url_val
            import io as _io2
            buf = _io2.BytesIO()
            wb.save(buf)
            data = buf.getvalue()
            self.send_file(data, "JAC_export.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        else:
            self.send_html("<h1>400 Unknown format</h1>", 400)

    def do_GET(self):
        # 检查认证
        if not self.require_auth():
            return

        parsed = urlparse(self.path)
        path   = parsed.path
        qs     = parse_qs(parsed.query)
        params = {k: v[0] for k, v in qs.items() if v}

        conn = get_conn()
        try:
            if path == "/":
                # Default: show all articles
                page = int(params.get("page", 1))
                per_page = 50
                offset = (page - 1) * per_page
                articles = query_articles(conn, {}, limit=per_page, offset=offset)
                total = count_articles(conn, {})
                self.send_html(render_search(conn, {}, articles, total, page, per_page))

            elif path == "/search":
                page = int(params.get("page", 1))
                per_page = 50
                offset = (page - 1) * per_page
                articles = query_articles(conn, params, limit=per_page, offset=offset)
                total = count_articles(conn, params)
                self.send_html(render_search(conn, params, articles, total, page, per_page))

            elif path == "/stats":
                self.send_html(render_stats(conn))

            elif path == "/api/articles":
                articles = query_articles(conn, params, limit=500)
                self.send_json([dict(a) for a in articles])

            elif path == "/api/stats":
                stats = get_stats(conn)
                self.send_json({
                    "total": stats["total"],
                    "by_year": dict(stats["by_year"]),
                    "by_type": dict(stats["by_type"]),
                    "by_area": [dict(r) for r in stats["by_area"]],
                })

            elif path == "/api/export-ris":
                self._handle_export_ris(conn, params)

            elif path == "/api/export":
                self._handle_export(conn, params)

            elif path == "/health":
                # 健康检查端点
                health_status = {
                    "status": "healthy",
                    "timestamp": datetime.now().isoformat(),
                    "version": "v0.4.2",
                    "database": "connected" if conn else "disconnected"
                }
                self.send_json(health_status)

            else:
                self.send_html("<h1>404</h1>", 404)
        finally:
            conn.close()


# ─── Main ─────────────────────────────────────────────────────

def main():
    if not DB_PATH.exists():
        logger.error(f"DB not found: {DB_PATH}")
        logger.info("Run: python src/fetch.py 2025")
        sys.exit(1)

    server = HTTPServer(("localhost", PORT), Handler)
    logger.info(f"JAC_Year v0.4.2")
    logger.info(f"DB: {DB_PATH}")
    logger.info(f"Running: http://localhost:{PORT}")
    logger.info("Ctrl+C to stop")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        logger.info("\nStopped.")


if __name__ == "__main__":
    main()
